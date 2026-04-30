"""Import/ETL system for RecordX.Finance Tally clone.

Supports:
- Excel template downloads & uploads (ledgers, groups, vouchers, stock items)
- Bank statement import (CSV/Excel)
- Tally XML import (vouchers)
- Validation & error reporting
"""
from datetime import datetime, date
from typing import List, Dict, Optional, Any, Tuple
from pathlib import Path
import io
import xml.etree.ElementTree as ET
from decimal import Decimal

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ImportValidationError(Exception):
    """Validation error during import."""
    pass


class ImportETL:
    """ETL operations for importing data into Tally system."""

    def __init__(self, db):
        """Initialize with database instance."""
        self.db = db

    # ==================== EXCEL TEMPLATES ====================

    def generate_ledger_template(self) -> bytes:
        """Generate Excel template for ledger import."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel templates")

        wb = Workbook()
        ws = wb.active
        ws.title = "Ledgers"

        # Headers
        headers = ["Name", "Group Name", "Opening Balance", "Notes"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        # Sample data
        ws.append(["ABC Suppliers", "Sundry Creditors", 50000, "Opening balance as of FY start"])
        ws.append(["XYZ Customer", "Sundry Debtors", 75000, ""])
        ws.append(["HDFC Bank Current", "Bank Accounts", 100000, ""])

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def generate_group_template(self) -> bytes:
        """Generate Excel template for group import."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel templates")

        wb = Workbook()
        ws = wb.active
        ws.title = "Groups"

        # Headers
        headers = ["Name", "Parent Group", "Type", "Notes"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        # Sample data
        ws.append(["Branch Expenses", "Indirect Expenses", "Expense", ""])
        ws.append(["Marketing Expenses", "Indirect Expenses", "Expense", ""])
        ws.append(["Trade Debtors", "Sundry Debtors", "Asset", ""])

        # Type options note
        ws.append([])
        ws.append(["Type Options:", "Asset, Liability, Revenue, Expense"])

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def generate_voucher_template(self, voucher_type: str = "Payment") -> bytes:
        """Generate Excel template for voucher import."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel templates")

        wb = Workbook()
        ws = wb.active
        ws.title = f"{voucher_type} Vouchers"

        # Headers
        headers = ["Date", "Voucher Number", "Ledger Name", "Debit", "Credit", "Narration"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        # Sample data
        ws.append(["2026-04-15", "PMT/001", "HDFC Bank", "", 10000, "Payment to ABC Suppliers"])
        ws.append(["2026-04-15", "PMT/001", "ABC Suppliers", 10000, "", "Payment to ABC Suppliers"])
        ws.append([])
        ws.append(["2026-04-16", "PMT/002", "HDFC Bank", "", 5000, "Rent payment"])
        ws.append(["2026-04-16", "PMT/002", "Rent Expense", 5000, "", "Rent payment"])

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def generate_stock_item_template(self) -> bytes:
        """Generate Excel template for stock item import."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel templates")

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Items"

        # Headers
        headers = ["Name", "Unit", "Opening Stock", "Rate", "Category", "HSN Code", "GST Rate"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        # Sample data
        ws.append(["Widget A", "Nos", 100, 500, "Finished Goods", "8471", 18])
        ws.append(["Raw Material X", "Kg", 500, 200, "Raw Materials", "3901", 12])

        # Column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    # ==================== EXCEL IMPORT ====================

    def import_ledgers_from_excel(self, file_content: bytes) -> Dict[str, Any]:
        """Import ledgers from Excel file."""
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas required for Excel import")

        try:
            df = pd.read_excel(io.BytesIO(file_content))
        except Exception as e:
            raise ImportValidationError(f"Failed to read Excel file: {str(e)}")

        # Validate columns
        required_cols = ["Name", "Group Name", "Opening Balance"]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise ImportValidationError(f"Missing required columns: {', '.join(missing_cols)}")

        results = {
            "total": len(df),
            "success": 0,
            "errors": [],
            "created_ids": []
        }

        with self.db.session() as session:
            for idx, row in df.iterrows():
                try:
                    # Skip empty rows
                    if pd.isna(row["Name"]):
                        continue

                    name = str(row["Name"]).strip()
                    group_name = str(row["Group Name"]).strip()
                    opening_balance = float(row["Opening Balance"]) if not pd.isna(row["Opening Balance"]) else 0.0

                    # Validate group exists
                    group = self.db.get_group_by_name(group_name)
                    if not group:
                        results["errors"].append({
                            "row": idx + 2,
                            "error": f"Group not found: {group_name}"
                        })
                        continue

                    # Create ledger
                    ledger = self.db.create_ledger(
                        name=name,
                        group_id=group.id,
                        opening_balance=opening_balance
                    )
                    results["success"] += 1
                    results["created_ids"].append(ledger.id)

                except Exception as e:
                    results["errors"].append({
                        "row": idx + 2,
                        "error": str(e)
                    })

        return results

    def import_vouchers_from_excel(self, file_content: bytes, voucher_type: str, company_id: int) -> Dict[str, Any]:
        """Import vouchers from Excel file."""
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas required for Excel import")

        try:
            df = pd.read_excel(io.BytesIO(file_content))
        except Exception as e:
            raise ImportValidationError(f"Failed to read Excel file: {str(e)}")

        # Validate columns
        required_cols = ["Date", "Voucher Number", "Ledger Name", "Debit", "Credit", "Narration"]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise ImportValidationError(f"Missing required columns: {', '.join(missing_cols)}")

        # Get voucher type
        vtype = self.db.get_voucher_type_by_name(voucher_type.capitalize())
        if not vtype:
            raise ImportValidationError(f"Invalid voucher type: {voucher_type}")

        results = {
            "total_vouchers": 0,
            "success": 0,
            "errors": [],
            "created_ids": []
        }

        # Group rows by voucher number
        voucher_groups = {}
        for idx, row in df.iterrows():
            if pd.isna(row["Voucher Number"]):
                continue
            voucher_num = str(row["Voucher Number"]).strip()
            if voucher_num not in voucher_groups:
                voucher_groups[voucher_num] = []
            voucher_groups[voucher_num].append((idx, row))

        results["total_vouchers"] = len(voucher_groups)

        # Process each voucher
        for voucher_num, rows in voucher_groups.items():
            try:
                # Get voucher date and narration from first row
                first_row = rows[0][1]
                voucher_date = pd.to_datetime(first_row["Date"]).date()
                narration = str(first_row["Narration"]) if not pd.isna(first_row["Narration"]) else ""

                # Build entries
                entries = []
                for idx, row in rows:
                    ledger_name = str(row["Ledger Name"]).strip()
                    debit = float(row["Debit"]) if not pd.isna(row["Debit"]) else 0.0
                    credit = float(row["Credit"]) if not pd.isna(row["Credit"]) else 0.0

                    # Validate ledger exists
                    ledger = self.db.get_ledger_by_name(ledger_name)
                    if not ledger:
                        raise ValueError(f"Ledger not found: {ledger_name}")

                    if debit > 0:
                        entries.append({
                            "ledger_id": ledger.id,
                            "amount": debit,
                            "is_debit": True
                        })
                    elif credit > 0:
                        entries.append({
                            "ledger_id": ledger.id,
                            "amount": credit,
                            "is_debit": False
                        })

                # Validate voucher balances
                total_debit = sum(e["amount"] for e in entries if e["is_debit"])
                total_credit = sum(e["amount"] for e in entries if not e["is_debit"])
                if abs(total_debit - total_credit) > 0.01:
                    raise ValueError(f"Voucher {voucher_num} doesn't balance: Dr={total_debit}, Cr={total_credit}")

                # Create voucher
                voucher = self.db.create_voucher(
                    voucher_type_id=vtype.id,
                    voucher_number=voucher_num,
                    date=voucher_date,
                    company_id=company_id,
                    narration=narration,
                    entries=entries
                )
                results["success"] += 1
                results["created_ids"].append(voucher.id)

            except Exception as e:
                results["errors"].append({
                    "voucher": voucher_num,
                    "error": str(e)
                })

        return results

    # ==================== BANK STATEMENT IMPORT ====================

    def import_bank_statement(
        self,
        file_content: bytes,
        ledger_id: int,
        file_format: str = "excel",
        column_mapping: Optional[Dict[str, str]] = None
    ) -> Dict[str, Any]:
        """Import bank statement from CSV/Excel.

        Args:
            file_content: File bytes
            ledger_id: Bank ledger ID
            file_format: 'excel' or 'csv'
            column_mapping: Map file columns to expected columns
                           e.g., {"Transaction Date": "date", "Dr": "debit", "Cr": "credit"}
        """
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas required for bank statement import")

        # Read file
        try:
            if file_format == "csv":
                df = pd.read_csv(io.BytesIO(file_content))
            else:
                df = pd.read_excel(io.BytesIO(file_content))
        except Exception as e:
            raise ImportValidationError(f"Failed to read file: {str(e)}")

        # Apply column mapping if provided
        if column_mapping:
            df = df.rename(columns=column_mapping)

        # Validate required columns
        required = ["date", "description", "debit", "credit"]
        missing = [col for col in required if col not in df.columns]
        if missing:
            raise ImportValidationError(f"Missing columns after mapping: {', '.join(missing)}")

        results = {
            "total": len(df),
            "imported": 0,
            "matched": 0,
            "unmatched": 0,
            "errors": [],
            "suggestions": []
        }

        statements = []
        for idx, row in df.iterrows():
            try:
                # Parse date
                txn_date = pd.to_datetime(row["date"]).date()
                description = str(row["description"]) if not pd.isna(row["description"]) else ""
                debit = float(row["debit"]) if not pd.isna(row["debit"]) else 0.0
                credit = float(row["credit"]) if not pd.isna(row["credit"]) else 0.0
                balance = float(row["balance"]) if "balance" in df.columns and not pd.isna(row["balance"]) else 0.0
                cheque_num = str(row["cheque_number"]) if "cheque_number" in df.columns and not pd.isna(row["cheque_number"]) else None
                reference = str(row["reference"]) if "reference" in df.columns and not pd.isna(row["reference"]) else None

                statements.append({
                    "date": txn_date,
                    "description": description,
                    "debit": debit,
                    "credit": credit,
                    "balance": balance,
                    "cheque_number": cheque_num,
                    "reference": reference
                })

            except Exception as e:
                results["errors"].append({
                    "row": idx + 2,
                    "error": str(e)
                })

        # Import statements using banking module
        from .banking import BankingOperations
        banking = BankingOperations(self.db)
        imported = banking.import_bank_statement(ledger_id, statements)
        results["imported"] = len(imported)

        # Auto-match transactions
        for stmt in imported:
            matched = self._auto_match_statement(stmt, ledger_id)
            if matched:
                results["matched"] += 1
            else:
                results["unmatched"] += 1
                # Generate suggestions
                suggestion = self._suggest_ledger_mapping(stmt)
                if suggestion:
                    results["suggestions"].append({
                        "statement_id": stmt.id,
                        "description": stmt.description,
                        "amount": stmt.debit if stmt.debit > 0 else stmt.credit,
                        "suggested_ledger": suggestion
                    })

        return results

    def _auto_match_statement(self, statement, ledger_id: int) -> bool:
        """Auto-match bank statement to existing voucher."""
        from .models import Voucher, LedgerEntry, BankStatement
        from sqlalchemy import select

        with self.db.session() as session:
            # Try to match by cheque number
            if statement.cheque_number:
                query = select(Voucher).join(Voucher.cheques).where(
                    Voucher.bank_ledger_id == ledger_id
                ).where(
                    Voucher.cheques.any(cheque_number=statement.cheque_number)
                )
                voucher = session.execute(query).scalar_one_or_none()
                if voucher:
                    stmt_obj = session.get(BankStatement, statement.id)
                    stmt_obj.matched_voucher_id = voucher.id
                    session.commit()
                    return True

            # Try to match by amount and date (within 3 days)
            amount = statement.debit if statement.debit > 0 else statement.credit
            from datetime import timedelta
            date_from = statement.transaction_date - timedelta(days=3)
            date_to = statement.transaction_date + timedelta(days=3)

            query = select(Voucher).where(
                Voucher.bank_ledger_id == ledger_id,
                Voucher.date >= date_from,
                Voucher.date <= date_to
            )
            vouchers = session.execute(query).scalars().all()

            for voucher in vouchers:
                # Calculate voucher amount
                bank_entries = [e for e in voucher.entries if e.ledger_id == ledger_id]
                for entry in bank_entries:
                    if abs(entry.amount - amount) < 0.01:
                        stmt_obj = session.get(BankStatement, statement.id)
                        stmt_obj.matched_voucher_id = voucher.id
                        session.commit()
                        return True

        return False

    def _suggest_ledger_mapping(self, statement) -> Optional[str]:
        """Suggest ledger for unmatched statement based on description."""
        description_lower = statement.description.lower()

        # Simple keyword-based suggestions
        keywords = {
            "salary": "Salary Expense",
            "rent": "Rent Expense",
            "electricity": "Electricity Expense",
            "telephone": "Telephone Expense",
            "interest": "Interest Paid",
            "commission": "Commission Expense",
            "professional": "Professional Fees",
            "tax": "Tax Payment",
            "neft": "Bank Charges",
            "rtgs": "Bank Charges",
            "charges": "Bank Charges"
        }

        for keyword, ledger in keywords.items():
            if keyword in description_lower:
                return ledger

        return None

    # ==================== XML IMPORT ====================

    def import_vouchers_from_xml(self, xml_content: str, company_id: int) -> Dict[str, Any]:
        """Import vouchers from Tally XML format.

        Supports TALLYMESSAGE structure:
        <ENVELOPE>
          <BODY>
            <IMPORTDATA>
              <REQUESTDATA>
                <TALLYMESSAGE>
                  <VOUCHER>...</VOUCHER>
                </TALLYMESSAGE>
              </REQUESTDATA>
            </IMPORTDATA>
          </BODY>
        </ENVELOPE>
        """
        try:
            root = ET.fromstring(xml_content)
        except ET.ParseError as e:
            raise ImportValidationError(f"Invalid XML: {str(e)}")

        results = {
            "total": 0,
            "success": 0,
            "errors": [],
            "created_ids": []
        }

        # Find all VOUCHER elements
        vouchers = root.findall(".//VOUCHER")
        results["total"] = len(vouchers)

        for voucher_elem in vouchers:
            try:
                voucher_data = self._parse_tally_voucher_xml(voucher_elem)
                voucher = self._create_voucher_from_xml_data(voucher_data, company_id)
                results["success"] += 1
                results["created_ids"].append(voucher.id)
            except Exception as e:
                voucher_num = voucher_elem.findtext("VOUCHERNUMBER", "Unknown")
                results["errors"].append({
                    "voucher": voucher_num,
                    "error": str(e)
                })

        return results

    def _parse_tally_voucher_xml(self, voucher_elem: ET.Element) -> Dict[str, Any]:
        """Parse Tally voucher XML element."""
        data = {
            "voucher_type": voucher_elem.findtext("VOUCHERTYPENAME", "").strip(),
            "voucher_number": voucher_elem.findtext("VOUCHERNUMBER", "").strip(),
            "date": voucher_elem.findtext("DATE", "").strip(),
            "narration": voucher_elem.findtext("NARRATION", "").strip(),
            "entries": []
        }

        # Parse ledger entries
        for ledger_elem in voucher_elem.findall(".//ALLLEDGERENTRIES.LIST"):
            ledger_name = ledger_elem.findtext("LEDGERNAME", "").strip()
            amount_str = ledger_elem.findtext("AMOUNT", "0").strip()
            is_deemed_positive = ledger_elem.findtext("ISDEEMEDPOSITIVE", "No").strip().upper() == "YES"

            # Parse amount (Tally uses negative for debit)
            amount = abs(float(amount_str))
            is_debit = float(amount_str) < 0

            data["entries"].append({
                "ledger_name": ledger_name,
                "amount": amount,
                "is_debit": is_debit
            })

        return data

    def _create_voucher_from_xml_data(self, data: Dict[str, Any], company_id: int):
        """Create voucher from parsed XML data."""
        # Get voucher type
        vtype = self.db.get_voucher_type_by_name(data["voucher_type"])
        if not vtype:
            raise ValueError(f"Unknown voucher type: {data['voucher_type']}")

        # Parse date (Tally format: YYYYMMDD)
        date_str = data["date"]
        if len(date_str) == 8:
            voucher_date = datetime.strptime(date_str, "%Y%m%d").date()
        else:
            voucher_date = datetime.strptime(date_str, "%Y-%m-%d").date()

        # Build entries with ledger IDs
        entries = []
        for entry_data in data["entries"]:
            ledger = self.db.get_ledger_by_name(entry_data["ledger_name"])
            if not ledger:
                raise ValueError(f"Ledger not found: {entry_data['ledger_name']}")

            entries.append({
                "ledger_id": ledger.id,
                "amount": entry_data["amount"],
                "is_debit": entry_data["is_debit"]
            })

        # Create voucher
        return self.db.create_voucher(
            voucher_type_id=vtype.id,
            voucher_number=data["voucher_number"],
            date=voucher_date,
            company_id=company_id,
            narration=data["narration"],
            entries=entries
        )

    # ==================== VALIDATION ====================

    def validate_import_data(self, data: List[Dict], data_type: str) -> Dict[str, Any]:
        """Validate import data before processing.

        Args:
            data: List of dictionaries with import data
            data_type: 'ledger', 'voucher', 'group', etc.

        Returns:
            Validation report with errors and warnings
        """
        report = {
            "valid": True,
            "errors": [],
            "warnings": [],
            "total_records": len(data)
        }

        validators = {
            "ledger": self._validate_ledger_data,
            "voucher": self._validate_voucher_data,
            "group": self._validate_group_data
        }

        validator = validators.get(data_type)
        if not validator:
            report["errors"].append(f"Unknown data type: {data_type}")
            report["valid"] = False
            return report

        for idx, record in enumerate(data):
            errors = validator(record)
            if errors:
                report["valid"] = False
                report["errors"].extend([
                    {"record": idx + 1, "error": err} for err in errors
                ])

        return report

    def _validate_ledger_data(self, record: Dict) -> List[str]:
        """Validate ledger record."""
        errors = []
        if not record.get("name"):
            errors.append("Name is required")
        if not record.get("group_name"):
            errors.append("Group name is required")
        if "opening_balance" in record:
            try:
                float(record["opening_balance"])
            except (ValueError, TypeError):
                errors.append("Opening balance must be a number")
        return errors

    def _validate_voucher_data(self, record: Dict) -> List[str]:
        """Validate voucher record."""
        errors = []
        if not record.get("voucher_type"):
            errors.append("Voucher type is required")
        if not record.get("voucher_number"):
            errors.append("Voucher number is required")
        if not record.get("date"):
            errors.append("Date is required")
        if not record.get("entries") or len(record["entries"]) < 2:
            errors.append("At least 2 ledger entries required")
        return errors

    def _validate_group_data(self, record: Dict) -> List[str]:
        """Validate group record."""
        errors = []
        if not record.get("name"):
            errors.append("Name is required")
        if record.get("type") and record["type"] not in ["Asset", "Liability", "Revenue", "Expense"]:
            errors.append("Invalid group type")
        return errors
