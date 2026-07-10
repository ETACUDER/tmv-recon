"""Rooftop Restaurant voucher pipeline — raw monthly files → Tally XML.

Reproduces the accountant's hotel-analogous model, in 3 stages:

  STAGE 1  Sale (Journal)   Dr  channel debtor    Cr  SALES UNDER COMPOSITION SCHEME
             Paytm  → CARD / UPI / PAYTM / G PAY [F&B]   (from settlement report)
             Cash   → SANDEEP SHARMA IMP A/C.            (daily remainder)
  STAGE 2  Receipt           Dr  INDIAN BANK        Cr  channel debtor (net settled)
  STAGE 3  Payment           Dr  expense/supplier   Cr  INDIAN BANK

Sources:
  * Sales Detail (EZee POS, .html)  — dated daily F&B turnover
  * Settlement report (Paytm, .xlsx) — splits each day's sales into Paytm vs cash
  * Bank statement (.xlsx)          — receipts (credits) + payments (debits)

Swiggy / Zomato are left for the accountant (skipped here, reported separately).
Composition dealer: no output GST, no bill-wise allocation. Voucher schema is
cloned verbatim from the restaurant's own Tally export (templates/voucher.tmpl).
"""
from __future__ import annotations

import difflib
import json
import re
import uuid
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape as xesc

import openpyxl

_HERE = Path(__file__).resolve().parent
ROOT = _HERE.parents[2]
_TMPL = (_HERE / "templates" / "voucher.tmpl").read_text(encoding="utf-8")          # Receipt/Payment
_TMPL_SALE = (_HERE / "templates" / "sales.tmpl").read_text(encoding="utf-8")       # Sales (invoice)
_TMPL_JOURNAL = (_HERE / "templates" / "journal.tmpl").read_text(encoding="utf-8")  # mode-routing Journal
_LEDGERS_JSON = ROOT / "data" / "recon" / "rooftop" / "ledgers.json"

# ----- Company + fixed ledgers (all verified present in Master.xml) -----
COMPANY = "TMV Rooftop Restaurant"
CMP_GSTIN = "08GFRPS2684G1ZM"
BANK_LEDGER = "INDIAN BANK A/C 8150353104"
COLLECTION_LEDGER = "CARD / UPI / PAYTM / G PAY [F&B]"   # Paytm (Sundry Debtor)
BUNDAL_LEDGER = "BUNDAL TECHNOLOGIES"                    # Dineout settles via Bundl
SWIGGY_LEDGER = "SWIGGY SCR"                             # Swiggy
ZOMATO_LEDGER = "ZOMATO"                                 # Zomato
IMPREST_LEDGER = "SANDEEP SHARMA IMP A/C."              # cash float (Loans & Advances)
SALES_LEDGER = "SALES UNDER COMPOSITION SCHEME"         # revenue (Sales Accounts)
SALES_DEBTOR = "SUNDRY DEBTORS RESTAURENT"              # generic sales receivable
SUSPENSE_LEDGER = "SUSPENSE"
FY_PREFIX = "25-26"                                    # fallback only; real FY derived per order date (fy_prefix)

# EZee 'Settlement Detail' Payment channel -> Tally debtor ledger (all in Master).
# UPI/Card ("bank") -> the F&B card debtor; Dineout + Swiggy settle via Bundl; Zomato its own.
# Built-in defaults; the accountant can override any of these from the config page
# (saved to rooftop_channel_ledgers.json in TMV_DATA_DIR).
CHANNEL_LEDGER: dict[str, str] = {
    "CASH": IMPREST_LEDGER,
    "UPI": COLLECTION_LEDGER,
    "CREDIT CARD": COLLECTION_LEDGER,
    "DEBIT CARD": COLLECTION_LEDGER,
    "CARD": COLLECTION_LEDGER,
    "PAYTM": COLLECTION_LEDGER,
    "DINEOUT": BUNDAL_LEDGER,
    "SWIGGY": BUNDAL_LEDGER,
    "ZOMATO": ZOMATO_LEDGER,
}
# Journal flag per channel. True  = book the sale to SUNDRY DEBTORS RESTAURENT (New Ref)
#                                   then a routing Journal Dr <channel ledger> / Cr the debtor.
#                          False = book the sale DIRECTLY to <channel ledger> (party = that
#                                   ledger), no Journal — used for the aggregators so the
#                                   party shows as BUNDAL/ZOMATO, not the generic debtor.
# Aggregators default OFF (sale only); every other channel defaults ON ("as is").
DIRECT_CHANNELS: set[str] = {"DINEOUT", "SWIGGY", "ZOMATO"}
import os as _os
_DATA_BASE = _os.environ.get("TMV_DATA_DIR") or str(ROOT / "data")
CHANNEL_OVERRIDES_PATH = Path(_DATA_BASE) / "recon" / "rooftop_channel_ledgers.json"


def channel_map() -> dict[str, dict]:
    """Built-in channel map merged with the accountant's saved overrides.

    Returns {CHANNEL: {"ledger": str, "journal": bool}}. Overrides may be a plain
    string (ledger only — back-compat) or an object {"ledger", "journal"}.
    """
    m = {k: {"ledger": v, "journal": k not in DIRECT_CHANNELS}
         for k, v in CHANNEL_LEDGER.items()}
    try:
        for k, v in json.loads(CHANNEL_OVERRIDES_PATH.read_text(encoding="utf-8")).items():
            key = k.strip().upper()
            cur = dict(m.get(key) or {"ledger": None, "journal": key not in DIRECT_CHANNELS})
            if isinstance(v, dict):
                if v.get("ledger"):
                    cur["ledger"] = v["ledger"]
                if "journal" in v:
                    cur["journal"] = bool(v["journal"])
            else:
                cur["ledger"] = v
            m[key] = cur
    except (OSError, ValueError):
        pass
    return m


def _cled(cmap: dict, ch) -> str | None:
    """Ledger for a channel name from a channel_map() dict (None if unknown)."""
    ci = cmap.get((ch or "").upper())
    return ci["ledger"] if ci else None


def fy_prefix(yyyymmdd: str) -> str:
    """Indian fiscal-year label (Apr–Mar) for a YYYYMMDD date.

    EZee resets its receipt numbers each fiscal year, so the bill number's FY
    prefix must follow the order's own date, not a fixed value:
      20260331 -> '25-26'   20260401 -> '26-27'
    """
    try:
        y, m = int(str(yyyymmdd)[:4]), int(str(yyyymmdd)[4:6])
        s = y if m >= 4 else y - 1
        return f"{s % 100:02d}-{(s + 1) % 100:02d}"
    except (ValueError, TypeError):
        return FY_PREFIX

_GUID_PREFIX = "029dfefd-5996-4e71-8914-ec5a8528c655"
_GUID_NS = uuid.UUID(_GUID_PREFIX)


# ----- helpers -----
def _f(x) -> float:
    try:
        return float(str(x).replace(",", "").strip().strip("'"))
    except (ValueError, AttributeError):
        return 0.0


def _tdate(x) -> str:
    if isinstance(x, datetime):
        return x.strftime("%Y%m%d")
    s = str(x).strip().strip("'")[:10]
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    return re.sub(r"[^0-9]", "", s)[:8]


def _is_date(x) -> bool:
    return bool(re.match(r"\d{4}-\d{2}-\d{2}", str(x)))


# ----- parsers -----
def parse_settlement_detail(path: str | Path) -> dict:
    """EZee 'Settlement Detail' (.html) → {REC_no: {channel, ledger, amount, date}}.

    Rows are grouped under a single-cell channel header (Cash / UPI / Credit Card /
    Debit Card / Dineout / Zomato / Swiggy). Each order row is
    [Order No, Receipt No, Payment Type, Payment Date, Description, _, Amount, Order Date].
    The channel comes from the nearest preceding header — this is the per-order
    payment mode (no fuzzy matching needed).
    """
    html = Path(path).read_text(encoding="utf-8-sig", errors="replace")
    rows = [
        [re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", c)).strip()
         for c in re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", tr, re.S | re.I)]
        for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S | re.I)
    ]
    cmap = channel_map()
    known = set(cmap)
    out: dict[str, dict] = {}
    cur = None
    for r in rows:
        ne = [c for c in r if c.strip()]
        oid = next((c for c in r if c.startswith("ORD-")), None)
        if oid:
            v = r[r.index(oid):]
            rec = next((c for c in v if c.startswith("REC-")), "")
            amt = _f(v[6]) if len(v) > 6 else 0.0
            date = _tdate(v[7]) if len(v) > 7 else ""
            if rec:
                ci = cmap.get((cur or "").upper())
                out[rec] = {"channel": cur, "ledger": ci["ledger"] if ci else None,
                            "journal": ci["journal"] if ci else True,
                            "amount": round(amt, 2), "date": date}
        elif len(ne) == 1 and ne[0].upper() in known:
            cur = ne[0]
    return out


def parse_sales_detail(path: str | Path) -> dict:
    """EZee 'Sales Detail' (.html) → {tally_date: {'net': float, 'orders': int}}.

    Rows are grouped under a dd-mm-yyyy day header; each order line's
    'Final Total' (net = price - discount) is the customer bill value.
    """
    html = Path(path).read_text(encoding="utf-8-sig", errors="replace")
    rows = [
        [re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", c)).strip()
         for c in re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", tr, re.S | re.I)]
        for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S | re.I)
    ]
    dre = re.compile(r"^(\d{2})-(\d{2})-(\d{4})$")
    daily: dict[str, dict] = defaultdict(lambda: {"net": 0.0, "orders": 0})
    cur = None
    for r in rows:
        dc = [c for c in r if dre.match(c)]
        if dc and not any(c.startswith("ORD-") for c in r):
            m = dre.match(dc[0])
            cur = f"{m.group(3)}{m.group(2)}{m.group(1)}"  # YYYYMMDD
            continue
        oid = next((c for c in r if c.startswith("ORD-")), None)
        if oid and cur:
            vals = r[r.index(oid):]
            net = _f(vals[7]) if len(vals) > 7 else 0.0   # Final Total column
            daily[cur]["net"] += net
            daily[cur]["orders"] += 1
    return {k: {"net": round(v["net"], 2), "orders": v["orders"]} for k, v in daily.items()}


def parse_sales_detail_orders(path: str | Path) -> list[dict]:
    """EZee 'Sales Detail' (.html) → per-order list [{date, rec, inv, net}]."""
    html = Path(path).read_text(encoding="utf-8-sig", errors="replace")
    rows = [
        [re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", c)).strip()
         for c in re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", tr, re.S | re.I)]
        for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S | re.I)
    ]
    dre = re.compile(r"^(\d{2})-(\d{2})-(\d{4})$")
    orders, cur = [], None
    for r in rows:
        dc = [c for c in r if dre.match(c)]
        if dc and not any(c.startswith("ORD-") for c in r):
            m = dre.match(dc[0])
            cur = f"{m.group(3)}{m.group(2)}{m.group(1)}"
            continue
        oid = next((c for c in r if c.startswith("ORD-")), None)
        if oid and cur:
            vals = r[r.index(oid):]
            rec = next((c for c in vals if c.startswith("REC-")), "")
            net = _f(vals[7]) if len(vals) > 7 else 0.0
            if net > 0.005:
                orders.append({"date": cur, "rec": rec, "inv": f"{fy_prefix(cur)}/{rec}",
                               "net": round(net, 2)})
    return orders


def parse_settlement_daily(path: str | Path) -> dict:
    """Paytm settlement report → {tally_date: gross customer-paid} by Transaction Date."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = [r for r in wb.active.iter_rows(values_only=True)]
    wb.close()
    daily: dict[str, float] = defaultdict(float)
    for r in rows[1:]:
        if r[0] and str(r[0]).strip().upper() != "TOTAL" and r[1]:
            daily[_tdate(r[0])] += _f(r[1])
    return {k: round(v, 2) for k, v in daily.items()}


def parse_bank_statement(path: str | Path) -> dict:
    """Indian Bank statement → {'credits': [...], 'debits': [...]} (excludes summary row)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
    wb.close()
    hi = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Txn Date")
    credits, debits = [], []
    for r in rows[hi + 1:]:
        if not r or not _is_date(r[0]):
            continue
        desc = str(r[1] or "").strip()
        deb, cred = _f(r[3]), _f(r[4])
        if cred > 0:
            credits.append({"date": _tdate(r[0]), "desc": desc, "amount": cred})
        if deb > 0:
            debits.append({"date": _tdate(r[0]), "desc": desc, "amount": deb})
    return {"credits": credits, "debits": debits}


# ----- routing / matching -----
def route_credit(desc: str) -> dict:
    """Bank CREDIT → collection ledger, by merchant name in the NEFT narration."""
    u = desc.upper()
    if "CASH DEP" in u:
        return {"ledger": IMPREST_LEDGER, "score": 1.0, "category": "cash-deposit"}
    if "PAYTM" in u:
        return {"ledger": COLLECTION_LEDGER, "score": 1.0, "category": "paytm"}
    if "SWIGGY" in u or "BUNDL" in u or "BUNDAL" in u:
        return {"ledger": BUNDAL_LEDGER, "score": 1.0, "category": "swiggy"}
    if "ZOMATO" in u or "ETERNAL" in u:
        return {"ledger": ZOMATO_LEDGER, "score": 1.0, "category": "zomato"}
    return {"ledger": SUSPENSE_LEDGER, "score": 0.0, "category": "UNMAPPED"}


_STOP = {"A", "C", "AC", "A/C", "SALARY", "THE", "AND", "&", "CO", "PVT", "LTD",
         "PTE", "ENTERPRISES", "TRADERS", "AGENCIES", "BROTHERS", "BRO", "SONS"}


def _norm_tokens(s: str) -> list[str]:
    return [t for t in re.split(r"[^A-Za-z0-9]+", s.upper()) if t and t not in _STOP]


def _tok_hit(q: str, lt: list[str]) -> bool:
    for l in lt:
        if q == l:
            return True
        if len(q) >= 4 and (l.startswith(q) or q.startswith(l)):
            return True
        if difflib.SequenceMatcher(None, q, l).ratio() >= 0.82:
            return True
    return False


def _extract_party(desc: str) -> str:
    m = re.search(r"CLGINWPR:\s*([^,]+)", desc, re.I)
    if m:
        return m.group(1).strip()
    m = re.match(r"\s*([A-Za-z][A-Za-z .]+?)\s*/\s*TRANSFER TO", desc, re.I)
    if m:
        return m.group(1).strip()
    if desc.upper().startswith("TRANSFER TO"):
        return desc.split("/")[0].replace("TRANSFER TO", "").strip()
    return desc.split("/")[0].strip()


def load_ledgers() -> list[dict]:
    return json.loads(_LEDGERS_JSON.read_text(encoding="utf-8"))


def match_ledger(desc: str, ledgers: list[dict]) -> dict:
    """Bank DEBIT narration → best expense/supplier/salary ledger. Rules first, then fuzzy."""
    u = desc.upper()
    if "ESIC" in u:
        return {"ledger": "ESIC [ALI HUSSAIN BOHRA]", "score": 1.0, "party": "ESIC", "category": "statutory"}
    if any(k in u for k in ("CHRG", "CHARGE", "GST @", "SMS ALERT", "AMB ", "CASH HANDLING", "COLLECTION CHG")):
        return {"ledger": "BANK CHARGES", "score": 1.0, "party": "BANK CHARGES", "category": "bank-charges"}
    party = _extract_party(desc)
    qt = _norm_tokens(party)
    best, best_score = None, 0.0
    if qt:
        for led in ledgers:
            lt = _norm_tokens(led["name"])
            if not lt:
                continue
            overlap = sum(1 for q in qt if _tok_hit(q, lt)) / len(qt)
            ratio = difflib.SequenceMatcher(None, party.upper(), led["name"].upper()).ratio()
            score = 0.75 * overlap + 0.25 * ratio
            if score > best_score:
                best, best_score = led["name"], score
    if best and best_score >= 0.6:
        cat = "salary" if "SALARY" in u else "supplier/expense"
        return {"ledger": best, "score": round(best_score, 2), "party": party, "category": cat}
    return {"ledger": SUSPENSE_LEDGER, "score": round(best_score, 2), "party": party, "category": "UNMAPPED"}


# ----- voucher rendering -----
def _guid(month: str, vtype: str, seq: int) -> str:
    return str(uuid.uuid5(_GUID_NS, f"rooftop:{month}:{vtype}:{seq}"))


def render_voucher(*, vtype: str, date: str, party: str, contra: str,
                   party_is_debit: bool, amount: float, narration: str,
                   vnum: int, seq: int, month: str, alter_base: int) -> str:
    """One voucher: leg1 = `party` (PARTYLEDGERNAME), leg2 = `contra`. `amount` absolute."""
    amt = f"{abs(amount):.2f}"
    neg = f"-{amt}"
    if party_is_debit:                 # Dr party / Cr contra  (Payment, Sale)
        e1_amt, e1_d, e2_amt, e2_d = neg, "Yes", amt, "No"
    else:                              # Cr party / Dr contra  (Receipt)
        e1_amt, e1_d, e2_amt, e2_d = amt, "No", neg, "Yes"
    g = _guid(month, vtype, seq)
    aid = alter_base + seq
    sub = {
        "REMOTEID": g, "VCHKEY": f"{_GUID_PREFIX}-{(0xc000 + seq):08x}:{seq:08x}",
        "VTYPE": vtype, "DATE": date, "GUID": g,
        "NARRATION": xesc(narration), "PARTY": xesc(party), "LEG2": xesc(contra),
        "VNUM": str(vnum), "ALTERID": str(aid), "MASTERID": str(aid),
        "VKEY": str(198178380000000 + aid), "VRETAIN": str(seq),
        "E1_AMT": e1_amt, "E1_DEEMED": e1_d, "E1_LASTDEEMED": e1_d,
        "E2_AMT": e2_amt, "E2_DEEMED": e2_d, "E2_LASTDEEMED": e2_d,
    }
    out = _TMPL
    for k, v in sub.items():
        out = out.replace("{" + k + "}", v)
    return out


def _fill(tmpl: str, sub: dict) -> str:
    out = tmpl
    for k, v in sub.items():
        out = out.replace("{" + k + "}", v)
    return out


def _ids(kind: str, month: str, seq: int, alter_base: int) -> dict:
    g = str(uuid.uuid5(_GUID_NS, f"rooftop:{month}:{kind}:{seq}"))
    aid = alter_base + seq
    return {"REMOTEID": g, "GUID": g,
            "VCHKEY": f"{_GUID_PREFIX}-{(0xb400 + seq):08x}:{seq:08x}",
            "ALTERID": str(aid), "MASTERID": str(aid),
            "VKEY": str(197916387900000 + aid), "VRETAIN": str(seq)}


def render_sale(*, inv: str, date: str, amount: float, mode: str, party: str,
                seq: int, month: str, alter_base: int) -> str:
    """Sales invoice: Dr `party` (New Ref) / Cr SALES UNDER COMPOSITION SCHEME.

    `party` is SUNDRY DEBTORS RESTAURENT when a routing Journal follows, or the
    channel's own ledger (BUNDAL/ZOMATO/…) when the sale is booked directly.
    """
    amt = f"{abs(amount):.2f}"
    sub = {**_ids("Sales", month, seq, alter_base), "DATE": date, "INV": xesc(inv),
           "PARTY": xesc(party),
           "NARR": xesc(f"I.NO {inv} AGAINST REC IN  {mode}"),
           "AMT_NEG": f"-{amt}", "AMT_POS": amt}
    return _fill(_TMPL_SALE, sub)


def render_journal(*, inv: str, date: str, amount: float, mode: str, channel: str,
                   seq: int, month: str, alter_base: int) -> str:
    """Mode routing: Dr channel debtor (New Ref) / Cr SUNDRY DEBTORS RESTAURENT (Agst Ref)."""
    amt = f"{abs(amount):.2f}"
    sub = {**_ids("Journal", month, seq, alter_base), "DATE": date, "INV": xesc(inv),
           "CHANNEL": xesc(channel), "NARR": xesc(f"I.NO {inv} AGAINST REC IN  {mode}"),
           "AMT_NEG": f"-{amt}", "AMT_POS": amt}
    return _fill(_TMPL_JOURNAL, sub)


def _wrap_envelope(vouchers: list[str]) -> str:
    msgs = "\n".join(
        f'    <TALLYMESSAGE xmlns:UDF="TallyUDF">\n{v}\n    </TALLYMESSAGE>' for v in vouchers)
    return (
        "<ENVELOPE>\n <HEADER>\n  <TALLYREQUEST>Import Data</TALLYREQUEST>\n </HEADER>\n"
        " <BODY>\n  <IMPORTDATA>\n   <REQUESTDESC>\n    <REPORTNAME>All Masters</REPORTNAME>\n"
        "    <STATICVARIABLES>\n"
        f"     <SVCURRENTCOMPANY>{xesc(COMPANY)}</SVCURRENTCOMPANY>\n"
        "    </STATICVARIABLES>\n   </REQUESTDESC>\n   <REQUESTDATA>\n"
        f"{msgs}\n   </REQUESTDATA>\n  </IMPORTDATA>\n </BODY>\n</ENVELOPE>\n")


def _write_xml(path: Path, envelope: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"\xff\xfe" + envelope.encode("utf-16-le"))


# ----- verification: our output vs the accountant's actual Tally export -----
def _load_vouchers(path: str | Path):
    import xml.etree.ElementTree as ET
    raw = Path(path).read_bytes().decode("utf-16")
    root = ET.fromstring(re.sub(r"&#[0-9]+;", " ", raw).encode("utf-8"))
    return root.findall(".//VOUCHER")


def _legs(v):
    out = []
    for c in ("ALLLEDGERENTRIES.LIST", "LEDGERENTRIES.LIST"):
        for e in v.findall("./" + c):
            out.append((e.findtext("LEDGERNAME"), float(e.findtext("AMOUNT") or 0)))
    return out


def compare_to_tally(gen_path, tally_path) -> dict:
    """Compare our generated XML against the accountant's actual Tally export."""
    G, R = _load_vouchers(gen_path), _load_vouchers(tally_path)
    gt = Counter(v.get("VCHTYPE") for v in G)
    rt = Counter(v.get("VCHTYPE") for v in R)
    notes = {"Sales": "Tally has a few extra (Zomato / adjustments)",
             "Journal": "extra = complimentary / owner / Zomato routing",
             "Receipt": "Tally includes Swiggy / Zomato collections we defer",
             "Payment": "same bank debits",
             "Purchase": "Swiggy/Zomato RCM commission + supplier bills — accountant only"}
    counts = [{"type": t, "ours": gt.get(t, 0), "tally": rt.get(t, 0),
               "diff": gt.get(t, 0) - rt.get(t, 0), "note": notes.get(t, "")}
              for t in ["Sales", "Journal", "Receipt", "Payment", "Purchase"]]

    def sales_map(V):
        m = {}
        for v in V:
            if v.get("VCHTYPE") == "Sales":
                m[v.findtext("VOUCHERNUMBER")] = round(max((a for _, a in _legs(v)), default=0), 2)
        return m
    gs, rs = sales_map(G), sales_map(R)
    common = set(gs) & set(rs)
    amt_ok = sum(1 for i in common if gs[i] == rs[i])
    sales = {"ours": len(gs), "tally": len(rs), "common": len(common), "amount_match": amt_ok,
             "match_pct": round(100 * amt_ok / max(len(common), 1)),
             "only_ours": len(set(gs) - set(rs)), "only_tally": len(set(rs) - set(gs)),
             "turnover_ours": round(sum(gs.values()), 2), "turnover_tally": round(sum(rs.values()), 2),
             "turnover_diff": round(sum(rs.values()) - sum(gs.values()), 2),
             # row-level issues
             "mismatch_rows": sorted(({"inv": i, "ours": gs[i], "tally": rs[i]}
                                      for i in common if gs[i] != rs[i]), key=lambda x: x["inv"]),
             "only_tally_rows": sorted(({"inv": i, "amount": rs[i]} for i in set(rs) - set(gs)),
                                       key=lambda x: x["inv"]),
             "only_ours_rows": sorted(({"inv": i, "amount": gs[i]} for i in set(gs) - set(rs)),
                                      key=lambda x: x["inv"])}

    def jchan(V):
        m = {}
        for v in V:
            if v.get("VCHTYPE") == "Journal":
                dr = [n for n, a in _legs(v) if a < 0]
                if dr:
                    m.setdefault(v.findtext("VOUCHERNUMBER"), dr[0])
        return m

    def nm(x):
        x = (x or "").upper()
        return "Paytm" if "CARD / UPI" in x else "Cash" if "SANDEEP" in x else "Other"
    def short(x):
        x = (x or "")
        return "CARD/UPI/PAYTM [F&B]" if "CARD / UPI" in x.upper() else \
               "SANDEEP IMP A/C." if "SANDEEP" in x.upper() else x
    gj, rj = jchan(G), jchan(R)
    cj = set(gj) & set(rj)
    agree = sum(1 for i in cj if nm(gj[i]) == nm(rj[i]))
    dis = Counter((nm(gj[i]), nm(rj[i])) for i in cj if nm(gj[i]) != nm(rj[i]))
    dis_rows = sorted(({"inv": i, "amount": gs.get(i, 0.0),
                        "ours": nm(gj[i]), "tally": nm(rj[i]), "tally_ledger": short(rj[i])}
                       for i in cj if nm(gj[i]) != nm(rj[i])),
                      key=lambda x: (x["ours"], -x["amount"]))
    mode = {"common": len(cj), "agree": agree, "disagree": len(cj) - agree,
            "agree_pct": round(100 * agree / max(len(cj), 1)),
            "breakdown": [{"ours": o, "tally": r, "count": n} for (o, r), n in dis.most_common()],
            "rows": dis_rows}

    total_match = round(100 * (len(G)) / max(len(R), 1))
    return {"counts": counts, "ours_total": len(G), "tally_total": len(R),
            "sales": sales, "mode": mode,
            "payments": {"ours": gt.get("Payment", 0), "tally": rt.get("Payment", 0)},
            "deferred": {"purchase": rt.get("Purchase", 0),
                         "sw_zo_receipts": rt.get("Receipt", 0) - gt.get("Receipt", 0)},
            "coverage_pct": total_match}


def _settlement_txns(path: str | Path) -> list[tuple]:
    """Paytm settlement → [(tally_date, amount)] per transaction (for per-order mode match)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = [r for r in wb.active.iter_rows(values_only=True)]
    wb.close()
    return [(_tdate(r[0]), round(_f(r[1]), 2)) for r in rows[1:]
            if r[0] and str(r[0]).strip().upper() != "TOTAL" and r[1]]


def generate(sales_html, settlement_detail_html, out_path, month: str,
             bank_path=None, alter_base: int = 300000) -> dict:
    """Per-order Sales (+ routing Journal) + optional bank Receipts/Payments → Tally XML.

    Per order (Sales Detail = amounts, Settlement Detail = payment channel). Each
    channel carries a `journal` flag (config page / rooftop_channel_ledgers.json):
      journal ON  (Cash/UPI/Card, default):
        Sales    Dr SUNDRY DEBTORS RESTAURENT (New Ref) / Cr SALES UNDER COMPOSITION SCHEME
        Journal  Dr <channel debtor> (New Ref)          / Cr SUNDRY DEBTORS RESTAURENT (Agst Ref)
      journal OFF (Dineout/Swiggy/Zomato aggregators, default):
        Sales    Dr <channel ledger> (New Ref)          / Cr SALES UNDER COMPOSITION SCHEME
        (booked directly to the aggregator ledger — no intermediary, no Journal)
    Channel comes straight from the Settlement Detail; no fuzzy matching. Bank
    statement is OPTIONAL — with it, Receipts (credits) + Payments (debits) are
    added; without it, only Sales (+ Journals) are generated.
    """
    ledgers = load_ledgers()
    cmap = channel_map()
    orders = parse_sales_detail_orders(sales_html)
    channels = parse_settlement_detail(settlement_detail_html)   # REC -> {channel, ledger, ...}

    by_channel: dict[str, list] = defaultdict(lambda: [0, 0.0])
    unrouted: list[dict] = []
    for o in orders:
        info = channels.get(o["rec"])
        if info and info.get("ledger"):
            o["mode"] = (info["channel"] or "").upper()
            o["channel"] = info["ledger"]
            o["channel_name"] = info["channel"]
            o["journal"] = info.get("journal", True)
        else:
            o["mode"], o["channel"] = "CASH", IMPREST_LEDGER
            o["channel_name"] = (info.get("channel") if info else None) or "(not in settlement)"
            o["journal"] = cmap.get("CASH", {}).get("journal", True)
            unrouted.append(o)
        by_channel[o["channel_name"]][0] += 1
        by_channel[o["channel_name"]][1] += o["net"]

    # ----- optional bank: Receipts (credits) + Payments (debits) -----
    receipts, payments, skipped = [], [], []
    bank_used = bool(bank_path)
    if bank_used:
        bank = parse_bank_statement(bank_path)
        for c in bank["credits"]:
            m = route_credit(c["desc"])
            if m["category"] in ("swiggy", "zomato"):
                skipped.append({**c, "category": m["category"]})
                continue
            receipts.append({**c, "ledger": m["ledger"], "kind": "receipt-" + m["category"], "map": m})
        for dbt in bank["debits"]:
            m = match_ledger(dbt["desc"], ledgers)
            payments.append({**dbt, "ledger": m["ledger"], "map": m})

    # ----- render -----
    # journal ON  → Sale to SUNDRY DEBTORS RESTAURENT + routing Journal to the channel ledger.
    # journal OFF → Sale booked DIRECTLY to the channel ledger (party = it), no Journal.
    vouchers, seq, journal_count = [], 0, 0
    for o in sorted(orders, key=lambda x: (x["date"], x["rec"])):
        use_journal = o.get("journal", True)
        party = SALES_DEBTOR if use_journal else o["channel"]
        seq += 1
        vouchers.append(render_sale(inv=o["inv"], date=o["date"], amount=o["net"],
                                    mode=o["mode"], party=party, seq=seq, month=month,
                                    alter_base=alter_base))
        if use_journal:
            seq += 1
            journal_count += 1
            vouchers.append(render_journal(inv=o["inv"], date=o["date"], amount=o["net"], mode=o["mode"],
                                           channel=o["channel"], seq=seq, month=month, alter_base=alter_base))
    for r in sorted(receipts, key=lambda x: x["date"]):
        seq += 1
        vouchers.append(render_voucher(vtype="Receipt", date=r["date"], party=r["ledger"],
                                       contra=BANK_LEDGER, party_is_debit=False, amount=r["amount"],
                                       narration=r["desc"], vnum=seq, seq=seq, month=month, alter_base=alter_base))
    for p in sorted(payments, key=lambda x: x["date"]):
        seq += 1
        vouchers.append(render_voucher(vtype="Payment", date=p["date"], party=p["ledger"],
                                       contra=BANK_LEDGER, party_is_debit=True, amount=p["amount"],
                                       narration=p["desc"], vnum=seq, seq=seq, month=month, alter_base=alter_base))
    _write_xml(Path(out_path), _wrap_envelope(vouchers))

    # ----- exceptions -----
    exceptions: list[dict] = []

    def _exc(category, severity, date, ref, amount, ledger, action):
        exceptions.append({"category": category, "severity": severity, "date": date,
                           "ref": ref, "amount": round(amount, 2), "ledger": ledger, "action": action})

    for o in unrouted:
        _exc("UNROUTED_ORDER", "high", o["date"], o["inv"], o["net"], IMPREST_LEDGER,
             f"Order not found in Settlement Detail (channel: {o['channel_name']}) — booked as CASH. "
             "Verify the payment channel and reclassify if needed.")
    for p in payments:
        cat, sc = p["map"]["category"], p["map"]["score"]
        if cat == "UNMAPPED":
            _exc("UNMAPPED_PAYMENT", "high", p["date"], p["desc"][:60], p["amount"], SUSPENSE_LEDGER,
                 "Bank payment narration matched no ledger — posted to SUSPENSE. Assign the correct ledger.")
        elif sc < 0.75:
            _exc("LOW_CONF_PAYMENT", "medium", p["date"], p["desc"][:60], p["amount"], p["ledger"],
                 f"Payment auto-mapped with low confidence ({sc}). Confirm the ledger.")
    for s in skipped:
        _exc("SWIGGY_ZOMATO", "info", s["date"], s["desc"][:60], s["amount"], s["category"],
             "Swiggy/Zomato bank credit — receipt left for the accountant.")

    sev_rank = {"high": 0, "medium": 1, "info": 2}
    exceptions.sort(key=lambda e: (sev_rank[e["severity"]], e["category"], e["date"]))
    exc_summary = dict(Counter(e["category"] for e in exceptions))

    # ----- reconciliation (sales by channel; + bank if provided) -----
    def _n(items):
        return len(items), round(sum(i["amount"] for i in items), 2)

    def row(label, count, amount, role="line"):
        return {"label": label, "count": count, "amount": round(amount, 2), "role": role}

    sales_total = round(sum(o["net"] for o in orders), 2)
    chan_rows = [row(f"{ch} → {_cled(cmap, ch) or '(review)'}", n, a)
                 for ch, (n, a) in sorted(by_channel.items(), key=lambda x: -x[1][1])]
    chan_rows.append(row("Total turnover (orders)", len(orders), sales_total, "total"))
    recon = [{"title": "F&B sales by channel (EZee Settlement Detail)", "rows": chan_rows}]
    if bank_used:
        recon.append({"title": "Bank collections — credits", "rows": [
            row("Booked as receipts", *_n(receipts)),
            row("Swiggy / Zomato → accountant", *_n(skipped), "pending"),
            row("Total bank credits", len(receipts) + len(skipped),
                round(sum(r["amount"] for r in receipts) + sum(s["amount"] for s in skipped), 2), "total")]})
        recon.append({"title": "Bank payments — debits", "rows": [
            row("Mapped to a ledger", *_n(payments), "total")]})

    sales_paytm = round(sum(a for ch, (n, a) in by_channel.items()
                            if _cled(cmap, ch) == COLLECTION_LEDGER), 2)
    sales_cash = round(sum(a for ch, (n, a) in by_channel.items()
                           if _cled(cmap, ch) == IMPREST_LEDGER), 2)
    return {
        "month": month, "out": str(out_path), "total_vouchers": len(vouchers),
        "orders": len(orders), "sales_count": len(orders), "journal_count": journal_count,
        "sales_total": sales_total,
        "sales_by_channel": {ch: {"orders": n, "amount": round(a, 2), "ledger": _cled(cmap, ch)}
                             for ch, (n, a) in by_channel.items()},
        "sales_paytm": sales_paytm, "sales_cash": sales_cash,
        "bank_used": bank_used,
        "receipts": len(receipts), "payments": len(payments),
        "receipt_total": round(sum(r["amount"] for r in receipts), 2),
        "payment_total": round(sum(p["amount"] for p in payments), 2),
        "swiggy_zomato_skipped": {"count": len(skipped),
                                  "amount": round(sum(s["amount"] for s in skipped), 2)},
        "unrouted_orders": len(unrouted),
        "unmapped_count": sum(1 for p in payments if p["map"]["category"] == "UNMAPPED"),
        "payment_map": [{"date": p["date"], "amount": p["amount"], "ledger": p["ledger"],
                         "score": p["map"]["score"], "category": p["map"]["category"],
                         "desc": p["desc"][:80]} for p in payments],
        "exceptions": exceptions, "exceptions_summary": exc_summary,
        "exception_count": len(exceptions),
        "exception_amount": round(sum(e["amount"] for e in exceptions), 2),
        "recon": recon,
    }
